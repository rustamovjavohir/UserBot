from datetime import datetime, date as datetime_date, timedelta

import pytz
from django.utils.timezone import is_aware, make_naive
import calendar
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, NamedStyle, numbers, DEFAULT_FONT, PatternFill, Color


def workers_2_xlsx2(query):
    HEADER_ROW = 1
    BODY_ROW = 2
    today = datetime.now()
    date = today.strftime("%Y-%m-%d")
    openpyxl.load_workbook(filename='media/check_in_out.xlsx')
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    bold_font = Font(bold=True)
    alignment = Alignment(horizontal='center', vertical='center')

    pk = query.model._meta.get_field('id')
    full_name = query.model._meta.get_field('full_name')
    department = query.model._meta.get_field('department')
    job = query.model._meta.get_field('job')
    role = query.model._meta.get_field('role')
    phone = query.model._meta.get_field('phone')

    worksheet.cell(row=HEADER_ROW, column=1, value=pk.verbose_name).font = bold_font
    worksheet.cell(row=HEADER_ROW, column=2, value=department.verbose_name).font = bold_font
    worksheet.cell(row=HEADER_ROW, column=3, value=full_name.verbose_name).font = bold_font
    worksheet.cell(row=HEADER_ROW, column=4, value=job.verbose_name).font = bold_font
    worksheet.cell(row=HEADER_ROW, column=5, value=role.verbose_name).style = bold_font, alignment

    for enum, data in enumerate(query, start=1):
        worksheet.cell(row=BODY_ROW, column=1, value=enum).alignment = alignment
        worksheet.cell(row=BODY_ROW, column=2, value=data.department.name).alignment = alignment
        worksheet.cell(row=BODY_ROW, column=3, value=data.full_name).alignment = alignment
        worksheet.cell(row=BODY_ROW, column=4, value=data.job).alignment = alignment
        BODY_ROW += 1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Workers-{date}.xlsx'
    workbook.save(response)
    return response


def workers_2_xlsx(query, start_date=None, end_date=None):
    month_style = NamedStyle(name='month_style', number_format='mmmm', font=Font(name='Calibri', size=11, bold=True))
    time_style = NamedStyle(name='time_style', number_format='h:mm')
    date_style = NamedStyle(name='date_style', number_format='dd/mm/yyyy')
    start_time_style = NamedStyle(name='start_time_style', number_format='h:mm', font=Font(bold=True))
    today = datetime.now()
    start_date = datetime.strptime(start_date, '%Y-%m-%d') if start_date else datetime_date(today.year, today.month, 1)
    end_date = datetime.strptime(end_date, '%Y-%m-%d') if end_date else datetime_date(today.year, today.month, 30)
    start_date = datetime_date(start_date.year, start_date.month, start_date.day)
    end_date = datetime_date(end_date.year, end_date.month, end_date.day)
    delta = end_date - start_date
    BODY_ROW = 3
    workbook = openpyxl.load_workbook(filename='media/check_in_out.xlsx')
    worksheet = workbook.active
    month_cell = worksheet.cell(row=1, column=1, value=start_date)
    month_cell.style = month_style
    month_cell.alignment = Alignment(horizontal='center', vertical='center')
    month_cell.number_format = numbers.FORMAT_DATE_TIME6
    worksheet.cell(row=1, column=2).style = start_time_style
    worksheet.cell(row=1, column=3).style = start_time_style
    for enum, data in enumerate(query, start=1):
        worksheet.cell(row=BODY_ROW, column=1, value=enum)
        worksheet.cell(row=BODY_ROW, column=2, value=data.department.name)
        worksheet.cell(row=BODY_ROW, column=3, value=data.full_name)
        for i in range(4, 2 * delta.days + 5):
            time_keeping_date = data.timekeeping_set.filter(date=timedelta(days=(i - 4) / 2) + start_date).first()
            try:
                if time_keeping_date:
                    if i % 2 == 0:
                        time_keeping_check_in = timedelta(hours=5) + time_keeping_date.check_in.replace(tzinfo=None)
                        check_in_cell = worksheet.cell(row=BODY_ROW, column=i, value=time_keeping_check_in)
                        check_in_cell.number_format = numbers.FORMAT_DATE_TIME6
                        check_in_cell.style = time_style
                        check_in_cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        time_keeping_check_out = timedelta(hours=5) + time_keeping_date.check_out.replace(tzinfo=None)
                        check_out_cell = worksheet.cell(row=BODY_ROW, column=i,
                                                        value=time_keeping_check_out)
                        check_out_cell.number_format = numbers.FORMAT_DATE_TIME6
                        check_out_cell.style = time_style
                        check_out_cell.alignment = Alignment(horizontal='center', vertical='center')
                if i % 2 == 0:
                    date_cell = worksheet.cell(row=2, column=i, value=timedelta(days=(i - 4) / 2) + start_date)
                    date_cell.style = date_style
                    date_cell.alignment = Alignment(horizontal='center', vertical='center')
            except Exception as ex:
                print(ex)
        BODY_ROW += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Workers-timekeeping.xlsx'
    workbook.save(response)
    return response


class WorkdayReportXlsx:

    def __init__(self, query, start_date, end_date):
        self.query = query
        self.start_date = start_date
        self.end_date = end_date
        self.workbook = openpyxl.load_workbook(filename='media/workday_report.xlsx')
        self.active_color = "92D050"
        self.inactive_color = "FF0000"
        self.day_color = "FFFF00"
        self.total_color = "5B9BD5"
        self.tz = pytz.timezone('Asia/Tashkent')

    @staticmethod
    def get_now():
        return datetime.now()

    def get_start_date(self):
        return datetime.strptime(self.start_date, '%Y-%m-%d') if self.start_date else datetime_date(
            datetime.now().year,
            datetime.now().month,
            1)

    def get_end_date(self):
        return datetime.strptime(self.end_date, '%Y-%m-%d') if self.end_date else datetime_date(
            datetime.now().year,
            datetime.now().month,
            31)

    def set_color(self, cell, color):
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def get_delta_day(self, delta):
        return self.get_start_date() + timedelta(days=delta)

    def get_delta(self):
        return self.get_end_date() - self.get_start_date()

    @property
    def get_worksheet(self):
        return self.workbook.active

    def set_now_date_value(self, row=1, column=1, value=datetime.now().strftime('%b %d')):
        return self.get_worksheet.cell(row=row, column=column, value=value)

    def set_now_date_style(self, cell):
        self.set_color(cell, self.day_color)

    def set_now_date(self, row, column, value):
        cell = self.set_now_date_value(row, column, value)
        self.set_now_date_style(cell)

    def set_workers_value(self, row, column, value):
        return self.get_worksheet.cell(row=row, column=column, value=value)

    def set_workers_style(self, cell):
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=11, bold=True, italic=True, underline="single")
        return cell

    def set_workers(self, row, column, value):
        cell = self.set_workers_value(row, column, value)
        self.set_workers_style(cell)

    def set_day_value(self, row, column, value):
        return self.get_worksheet.cell(row=row, column=column, value=value)

    def set_day_style(self, cell):
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=11, bold=True)
        self.set_color(cell, self.day_color)
        return cell

    def set_day(self, row, column, value):
        cell = self.set_day_value(row, column, value)
        self.set_day_style(cell)

    def set_work_time_value(self, row, column, value):
        return self.get_worksheet.cell(row=row, column=column, value=value)

    def set_work_time_style(self, cell):
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=11, bold=True)
        self.set_color(cell, self.inactive_color)
        return cell

    def set_work_time(self, row, column, value):
        cell = self.set_work_time_value(row, column, value)
        self.set_work_time_style(cell)
        if value:
            self.set_color(cell, self.active_color)

    def set_total_work_time_value(self, row, column, value):
        return self.get_worksheet.cell(row=row, column=column, value=value)

    def set_total_work_time_style(self, cell):
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=11, bold=True)

    def set_total_work_time(self, row, column, value):
        cell_title = self.set_total_work_time_value(1, column, "КОЛ-ВО")
        self.set_color(cell_title, self.day_color)
        cell = self.set_total_work_time_value(row, column, value)
        self.set_total_work_time_style(cell)
        self.set_color(cell, self.total_color)

    def calculate_daily_work_time(self, timekeeping):
        delta = timekeeping.work_time()
        minutes = getattr(delta, "seconds", 0) // 60
        if minutes < 450:
            return 0
        elif 470 < minutes < 490:
            return 1
        elif 490 < minutes < 510:
            return 1.2
        else:
            return 2

    def report(self):
        self.set_now_date(1, 1, datetime.now())
        for i in range(1, self.get_delta().days + 2):
            self.set_day(1, i + 1, self.get_delta_day(i - 1).day)
        for enum, data in enumerate(self.query, start=1):
            _daly_work_list = []
            self.set_workers(enum + 1, 1, data.full_name)
            timekeeping = data.timekeeping_set.all().filter(
                date__gte=self.get_start_date(),
                date__lte=self.get_end_date())
            if timekeeping:
                for i in range(1, self.get_delta().days + 2):
                    timekeeping_data = timekeeping.filter(date=self.get_start_date() + timedelta(days=i)).first()
                    if timekeeping_data:
                        _d = self.calculate_daily_work_time(timekeeping_data)
                        _daly_work_list.append(_d)
                        self.set_work_time(enum + 1, i + 2, _d)
                    else:
                        self.set_work_time(enum + 1, i + 2, None)
            self.set_total_work_time(enum + 1, self.get_delta().days + 3, sum(_daly_work_list))
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=workers.xlsx'
        self.workbook.save(response)
        return response
